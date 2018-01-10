import csv
import openpyxl
from tkinter import filedialog
from random import *

def ask_input():
    data = input("[1]Manual Input\n[2]Browse Data\n")

    if int(data) == 2:
        browse_data()
    else:
        hm_guest = input("How many guests?")
        print ("===SUGGESTED # of tables===")
        for i in range(1,int(hm_guest)+1):
            if int(hm_guest) % i == 0:
                print(i)

        hm_table = input("Enter # of tables: ")

        inp_tables = int(hm_guest) / int(hm_table)
        if inp_tables % 1 == 0:
            number_guest_table = int(hm_guest)/int(hm_table)
            print("Guest per table :",int(number_guest_table))
        else:
            print("fail")
            
        ask_for_guest(int(hm_guest))

def browse_data():
    o_csv = openpyxl.load_workbook('Matrix_Relationship_M.xlsm')
    s_o_csv = open("Simplified.csv","w")
    sheetname = o_csv.get_sheet_by_name('Sheet1')
    
    fname = filedialog.askopenfilename(filetypes = (("All files", "*.type"), ("All files", "*")))
    i=0
    with open(fname,"r") as file:
        for line in file:
            line = line.replace("\n","")
            #line = line.rstrip('\n')
            #print(i+1,line)
            sheetname.cell(row=1,column=i+2).value = line
            sheetname.cell(row=i+2,column=1).value = line
            print(line)
            i+=1
            

    tot_guest = i
    print ("===SUGGESTED # of tables===")
    for i in range(1,int(tot_guest)+1):
        if int(tot_guest) % i == 0:
            print(i)

    hm_table = input("Enter # of tables: ")

    inp_tables = int(tot_guest) / int(hm_table)
    if inp_tables % 1 == 0:
        number_guest_table = int(tot_guest)/int(hm_table)
        print("Guest per table :",int(number_guest_table))
    else:
        print("fail")
            
    for i in range (1,tot_guest+1):
        for j in range (2,tot_guest+1):
            row = sheetname.cell(row=i+1,column=1).value
            col = sheetname.cell(row=1,column=j+1).value
            if i+1 != j+1 and sheetname.cell(row=j+1,column=i+1).value is None:
                #print ("Relationship between",row,"and",col,":")
                #rel = input()

                #if str(rel) == "":
                #    rel = 0
                rel = randint(0,4)
                    
                merge = int(rel)
                s_merge = row,col,rel

                conv = str(s_merge)

                b = "()'"
                for char in b:
                    conv = conv.replace(char,"")
                    
                s_o_csv.write(conv)
                s_o_csv.write('\n')

                sheetname.cell(row=i+1,column=j+1).value = merge
                sheetname.cell(row=j+1,column=i+1).value = merge
                
    o_csv.save('Matrix_Relationship_SX.xlsx')
    s_o_csv.close()
    xlsx_csv()
    
def ask_for_guest(hm):
    o_csv = openpyxl.load_workbook('Matrix_Relationship_M.xlsm')
    s_o_csv = open("Simplified.csv","w")
    sheetname = o_csv.get_sheet_by_name('Sheet1')

    for i in range (hm):
        print (i+1,"guest name :")
        add_to_list =  input()
        sheetname.cell(row=1,column=i+2).value = add_to_list
        sheetname.cell(row=i+2,column=1).value = add_to_list

    for i in range (1,hm+1):
        for j in range (2,hm+1):
            row = sheetname.cell(row=i+1,column=1).value
            col = sheetname.cell(row=1,column=j+1).value
            if i+1 != j+1 and sheetname.cell(row=j+1,column=i+1).value is None:
                #print ("Relationship between",row,"and",col,":")
                #rel = input()

                #if str(rel) == "":
                #    rel = 0
                rel = randint(0,4)
                    
                merge = int(rel)
                s_merge = row,col,rel

                

                conv = str(s_merge)

                b = "()'"
                for char in b:
                    conv = conv.replace(char,"")
                    
                s_o_csv.write(conv)
                s_o_csv.write('\n')


                sheetname.cell(row=i+1,column=j+1).value = merge
                sheetname.cell(row=j+1,column=i+1).value = merge
                
    o_csv.save('Matrix_Relationship_SX.xlsx')
    s_o_csv.close()
    xlsx_csv()

def xlsx_csv():
    wb = openpyxl.load_workbook('Matrix_Relationship_SX.xlsx')
    sh = wb.get_active_sheet()
    with open('Matrix_Relationship_CSV.csv','w',newline='') as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    print("Converting done")
       
ask_input()
