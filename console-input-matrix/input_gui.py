import wx
import wx.grid as grid
import csv
import input_tabbing
import openpyxl
import math
import itertools
from tkinter import filedialog
from random import *
from time import gmtime, strftime
from tqdm import tqdm
#from input_gui import Frame

guest = []
currentgrid = 19
global totalcount
global tot_guest
global number_guest_table



class Frame(wx.Frame):
    def __init__(self, parent):
        wx.Frame.__init__(self, parent, -1, "Show Data", size=(700,450))
        
        menubar = wx.MenuBar()
        fileMenu = wx.Menu()
        fitem = fileMenu.Append(wx.ID_OPEN, '&Open')
        fitem1 = fileMenu.Append(wx.ID_NEW, '&New Data')
        fitem2 = fileMenu.Append(wx.ID_ANY, '&Input Data')
        fitem3 = fileMenu.Append(wx.ID_ANY, '&Process Data')

        menubar.Append(fileMenu, '&File')
        self.Bind(wx.EVT_MENU, self.browse_file, fitem)
        self.Bind(wx.EVT_MENU, self.new_data, fitem1)
        self.Bind(wx.EVT_MENU, self.input_data, fitem2)
        self.Bind(wx.EVT_MENU, self.browse_data, fitem3)
        self.SetMenuBar(menubar)
        
        self.grid = grid.Grid(self)
        self.grid.CreateGrid(currentgrid, currentgrid)
        self.grid.Bind(grid.EVT_GRID_CELL_CHANGED, self.OnCellChange)

    def OnCellChange(self, evt):
        print("OnCellChange: (%d,%d) %s\n" % (evt.GetRow(), evt.GetCol(), evt.GetPosition()))

        row = evt.GetRow()
        col = evt.GetCol()
        val = self.grid.GetCellValue(row, col)
        print(val)
        self.grid.SetCellValue(col,row,val)

    def input_data(self,e):
        test = input_tabbing.MainFrame()
        test.Show()
        print("test")

    def new_data(self,e):
        
        dlg = wx.TextEntryDialog(frame, 'How many guests?','Text Entry')
        #dlg.SetValue("Default")
        if dlg.ShowModal() == wx.ID_OK:
            print('You entered: %s\n' % dlg.GetValue())
        dlg.Destroy()

        totalcount = int(dlg.GetValue())+1
        
        self.grid.AppendRows(totalcount-currentgrid)
        self.grid.AppendCols(totalcount-currentgrid)
        self.grid.ClearGrid()
        #self.Layout()

        attr = wx.grid.GridCellAttr()
        attr.SetReadOnly(True)
        a=0
        for col in range(totalcount):
            for row in range(a,totalcount):
                self.grid.SetAttr(row,col,attr)
                self.grid.SetCellBackgroundColour(row,col,"gray")
            a+=1
        
        
    def browse_file(self,e):
        global totalcount
        with wx.FileDialog(self, "Open CSV file", wildcard="CSV file (*.csv)|*.csv",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return     # the user changed their mind
            
            pathname = fileDialog.GetPath()
            try:
                with open(pathname,newline="") as file:
                    row_count = sum(1 for row in file)
                    print(row_count)
                    totalcount = row_count
                    self.grid.AppendRows(row_count-currentgrid)
                    self.grid.AppendCols(row_count-currentgrid)
            except IOError:
                wx.LogError("Cannot open file '%s'." % newfile)
                

        self.display_data(pathname)

    def display_data(self,pathname):
        with open(pathname,newline="") as file:
            reader = csv.reader(file)
            r = 0
            for col in reader:
               c = 0
               for row in col:
                  self.grid.SetCellValue(r,c,row)
                  self.grid.SetCellValue(c,r,row)
                  if row == "":
                      self.grid.SetCellBackgroundColour(r,c,"gray")
                  c += 1 
               r += 1
               
    def browse_data(self,e):
        o_csv = openpyxl.load_workbook('Matrix_Relationship_M.xlsm')
        s_o_csv = open("Simplified.csv","w")
        sheetname = o_csv.get_sheet_by_name('Sheet1')
        
        with wx.FileDialog(self, "Open CSV file", wildcard="CSV file (*.txt)|*.txt",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return     # the user changed their mind
            
            pathname = fileDialog.GetPath()
        i=0
        with open(pathname,"r") as file:
            for line in file:
                line = line.replace("\n","")
                #line = line.rstrip('\n')
                #print(i+1,line)
                sheetname.cell(row=1,column=i+2).value = line
                sheetname.cell(row=i+2,column=1).value = line
                print(line)
                guest.append(line)
                i+=1
                
        global tot_guest
        tot_guest = i
        #print(tot_guest)
        #print(guest)
        print ("===SUGGESTED # of tables===")
        for i in range(1,int(tot_guest)+1):
            if int(tot_guest) % i == 0:
                print(i)

        dlg = wx.TextEntryDialog(frame, '# of tables?','Text Entry')
        if dlg.ShowModal() == wx.ID_OK:
            print('You entered: %s\n' % dlg.GetValue())
        dlg.Destroy()

        hm_table = dlg.GetValue()

        inp_tables = int(tot_guest) / int(hm_table)
        global number_guest_table
        if inp_tables % 1 == 0:
            number_guest_table = int(tot_guest)/int(hm_table)
            print("Guest per table :",int(number_guest_table))
        else:
            print("fail")

        
        c = 1
        
        for i in range (1,tot_guest+1):
            for j in range (2,tot_guest+1):
                row = sheetname.cell(row=i+1,column=1).value
                col = sheetname.cell(row=1,column=j+1).value
                if i+1 != j+1 and sheetname.cell(row=j+1,column=i+1).value is None:
##                    print("[0]No relationship")
##                    print("[100]Friends")
##                    print("[300]Aunt/Niece")
##                    print("[500]Cousin")
##                    print("[700]Parent/child")
##                    print("[900]Sibling")
##                    print("[2000]Spouse Date")
                    text = "Relationship between",row,"and",col,":"
                    text = str(text)
                    b = "()',"
                    for char in b:
                        text = text.replace(char,"")

                    print(tot_guest)
                    
                    tot_comb = ((int(tot_guest)-1) * (int(tot_guest)))/2
                    text_count = c,"/",int(tot_comb)
                    text_count = str(text_count)
                    for char in b:
                        text_count = text_count.replace(char,"")
    
                    dlg1 = wx.TextEntryDialog(frame,text ,text_count)
                    if dlg1.ShowModal() == wx.ID_OK:
                        print('You entered: %s\n' % dlg1.GetValue())
                    dlg1.Destroy()
                    
                    rel = dlg1.GetValue()

                    if str(rel) == "":
                        rel = 0
    ##                rel = randint(0,4)
    ##                    
                    merge = int(rel)
                    s_merge = row,col,rel
                    conv = str(s_merge)
                    
                    b = "()' "
                    for char in b:
                        conv = conv.replace(char,"")
                        
                    s_o_csv.write(conv)
                    s_o_csv.write('\n')

                    sheetname.cell(row=i+1,column=j+1).value = merge
                    sheetname.cell(row=j+1,column=i+1).value = merge
                    c+=1
                    
        o_csv.save('Matrix_Relationship_SX.xlsx')
        s_o_csv.close()
        self.xlsx_csv()
        
    
    def xlsx_csv(self):
        wb = openpyxl.load_workbook('Matrix_Relationship_SX.xlsx')
        sh = wb.get_active_sheet()
        with open('Matrix_Relationship_CSV.csv','w',newline='') as f:
            c = csv.writer(f)
            for r in sh.rows:
                c.writerow([cell.value for cell in r])
        print("Converting done")
        self.do_genetic('Simplified.csv')
        
    def do_genetic(self,filename):
        fitness = []
        countCrossover = 0
        crossoverMutate = []
        crossCount = 0
        
        startTime = strftime("%H:%M:%S", gmtime())
        crossoverCombination = list(itertools.combinations(guest, int(number_guest_table)))
        print(int(number_guest_table))
        with open(filename) as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            x = 0
            for row in readCSV:
                fitness.append([])
                fitness[x].append(row[0])
                fitness[x].append(row[1])
                fitness[x].append(row[2])
                x = x + 1
        #for h in fitness:
           # print(h)

           
        #Calculates the mutation per crossovers
        crossoverMutate.append([])
        for load in tqdm(range(len(crossoverCombination))):
            crossoverMutation = list(itertools.permutations(crossoverCombination[load]))
            count = 1
            maximum = 0
            mute = 0
            highestSet = []
            highestTotals = []

            #Getting the all possible set(5! = 120(Mutation))
            for x in crossoverMutation:
                pair = []
                mutate = x
                mutation = []
                mutationPair = []
                counting = 0
                total = 0
                check = True
                highestTotal = []

                #Getting the first 5 set posible pair
                for y in range(len(mutate)):
                    pair.append([])
                    first = y + len(mutate) - 1
                    if first > len(mutate)-1:
                        first = first - (len(mutate))
                    second = y
                    pair[y].append(mutate[first])
                    pair[y].append(mutate[second])
                #Getting the second 5 set posible pair
                for e in range(len(mutate)):
                    pair.append([])
                    focus = e - 1
                    if focus < 0:
                        focus = len(mutate)-1
                    first = focus - 1
                    if first < 0:
                        first = len(mutate) - 1
                    second = focus + 1
                    if second > len(mutate)-1:
                        second = 0
                    pair[len(mutate) + e].append(mutate[first])
                    pair[len(mutate) + e].append(mutate[second])

                #Getting its Fitness Value of the pair
                for a in range(len(pair)):
                    for b in range(len(fitness)):
                        if pair[a][0] == fitness[b][0] and pair[a][1] == fitness[b][1]:
                            mutation.append(fitness[b][2])
                        if pair[a][0] == fitness[b][1] and pair[a][1] == fitness[b][0]:
                            mutation.append(fitness[b][2])
                            
                #Calculating the first pair like (1+2)-(3+4)+(5+6)-(7+8)
                for c in range(int(len(mutation)/2)):
                    first = mutation[c + counting]
                    counting = counting+1
                    second = mutation[c + counting]
                    highestTotal.append(first)
                    highestTotal.append(second)
                    mutationPair.append(int(first) + int(second))

                #Calculating the whole Fitness Value like 3 - 7 + 11 - 15
                for d in range(len(mutationPair)):
                    if check:
                        total = total + mutationPair[d]
                        check = False
                    else:
                        total = total - mutationPair[d]
                        check = True
                        
                #print("Mutation", count," : ",pair, " Fitness : ", total)
                if maximum < total:
                    maximum = total
                    mute = count
                    highestSet = pair
                    highestTotals = highestTotal
                    
                count = count+1
            num = 0
            for n in highestSet:
                if num < int(numGuest):
                    crossoverMutate[crossCount].append(n[0])
                else:
                    break
                #print(n[0],"-",n[1]," => ",highestTotals[num])
                num = num + 1
            crossoverMutate[crossCount].append(maximum)
            crossoverMutate.append([])
            crossCount = crossCount + 1

        print("Time Start:", startTime)
        print("Time End:", strftime("%H:%M:%S", gmtime()))

if __name__ == '__main__':
    app = wx.App(False)
    frame = Frame(None)
    frame.Show()
    app.MainLoop()
